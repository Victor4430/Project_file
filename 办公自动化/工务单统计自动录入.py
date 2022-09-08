#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :py_venv -> 工务单统计自动录入.py
# @IDE       :PyCharm
# @Time      :2022/9/8 15:06
# @Author    :杨晓东
# @Email     :lzj155@foxmail.com
# @homepage  :www.demo520.com

import csv
import requests
from lxml import etree
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
import time

# 装饰器 计时器
def time_out(a_func):
    def clocked(*args, **kwargs):
        start = time.time()
        result = a_func(*args, **kwargs)
        end = time.time()
        print("程序：" + a_func.__name__,"    运行时间：" + str(end - start))
        return result
    return clocked



def run():
    # 设置开始网址
    url = "https://www.baidu.com"
    # 设置代理端口
    proxy = '127.0.0.1:4780'
    proxies = {"http": "http://" + proxy, "https": "http://" + proxy}
    # 设置请求头
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/99.0.4844.84 Safari/537.36"}
    # 开始请求网页
    req = requests.get(url, headers=headers, proxies=proxies)
    req.encoding = 'utf-8'
    print(req.text)

    # 解析网页
    html = etree.HTML(req.text)
    test = html.xpath('//*[@id="su"]/@value')[0]
    print(test, '访问成功！')

    # ip测试_显示当前ip地址
    ip_url = 'https://icanhazip.com'
    ip_req = requests.get(ip_url, headers=headers, proxies=proxies)
    print('当前ip：', ip_req.text)


# csv 保存_使用数组方法
def csv_data_save_1():
    # 传入一个文件对象，然后才能在这个文件对象的基础上调用csv的写入方法writerow（写入一行）writerrow（写入多行）
    # 定义标题栏
    headers = ['class', 'name', 'sex', 'height', 'year']
    # 内容_列表（数组）方式
    rows = [
        [1, 'xiaoming', 'male', 168, 23],
        [1, 'xiaohong', 'female', 162, 22],
        [2, 'xiaozhang', 'female', 163, 21],
        [2, 'xiaoli', 'male', 158, 21]
    ]
    # 如果打开csv文件出现空行的情况，那么需要添加一个参数 newline=''
    with open('test.csv', 'w', newline='') as f:
        # 创建csv操作句柄
        f_csv = csv.writer(f)
        # 利用句柄写入标题
        f_csv.writerow(headers)
        # 通过句柄写入内容
        f_csv.writerows(rows)


# csv 保存_使用字典方法
def csv_data_save_2():
    # 写入字典序列类型数据的时候，需要传入两个参数，
    # 一个是文件对象——f，
    # 一个是字段名称——fieldnames，
    # 到时候要写入表头的时候，只需要调用writerheader方法，
    # 写入一行字典系列数据调用writerrow方法，并传入相应字典参数，写入多行调用writerows

    # 定义标题栏
    headers = ['class', 'name', 'sex', 'height', 'year']
    # 内容_字典（序列）方式
    rows = [
        {'class': 1, 'name': 'xiaoming', 'sex': 'male', 'height': 168, 'year': 23},
        {'class': 1, 'name': 'xiaohong', 'sex': 'female', 'height': 162, 'year': 22},
        {'class': 2, 'name': 'xiaozhang', 'sex': 'female', 'height': 163, 'year': 21},
        {'class': 2, 'name': 'xiaoli', 'sex': 'male', 'height': 158, 'year': 21},
    ]
    # 如果打开csv文件出现空行的情况，那么需要添加一个参数 newline=''
    with open('./test.csv', 'w', newline='') as f:
        # 创建csv操作句柄 and 写入标题
        f_csv = csv.DictWriter(f, headers)
        # 调用 writerheader() 写入表头
        f_csv.writeheader()
        # 通过句柄写入内容
        f_csv.writerows(rows)


# csv 读取_方法
def csv_data_read():
    # 读取csv时需要使用reader，并传如一个文件对象，而且reader返回的是一个可迭代的对象，需要使用for循环遍历
    # 在上面，row是一个列表，如果想要查看固定的某列，则需要加上下标，例如我想要查看name，那么只需要改为row[1]
    with open('./test.csv') as f:
        f_csv = csv.reader(f)
        for row in f_csv:
            print(row)
            print(type(row))
            print(row[1])
            print(type(row[1]))

@time_out
def auto():
    # 提示手动输入开始的工作令编号
    start_num = input('请输入结束单号：')
    print('公务单自动录入程序  启动......')

    # 加载公务工作令    表1：
    wb_1 = openpyxl.load_workbook("./123.xlsx")
    print('正在加载公务工作令.....')

    # 获取所有工作簿名称[]，并打印
    sheets = wb_1.sheetnames

    # 将工作令序号倒序
    sheets = sheets[::-1]

    # 打印出倒序的工作令
    print(sheets)

    # 加载需要写入的工作簿      表2
    wb_2 = openpyxl.load_workbook('456.xlsx')
    print('正在加载工务单统计.....')

    # 定义表2句柄： 需要写入的工作表
    ws_2 = wb_2['9月份']

    # 修改工务单统计表的标题为‘9月份_动力部工务单统计_(8月21-9月20)’
    ws_2.cell(row=1, column=1).value = '9月份_动力部工务单统计_(8月21-9月20)'

    # 行计数器
    row = 0
    # 序号计数器
    num = 0

    # 遍历所有的工作表
    for i in sheets:
        # 序号递增
        num += 1

        # 表1句柄： 读取工作表操作：
        ws_1 = wb_1[i]
        print('工作簿编号： ', i)

        # 判断：  跳过  小余  输入的工作表编号
        if int(i) - int(start_num) < 0:
            print('跳过：  ', i)
            # 序号递减
            num -= 1
            # 跳出本次循环
            continue

        # 序号
        print('序号:', num)
        # 表2写入序号
        ws_2.cell(row=3 + row, column=1).value = num

        # 所在部门
        suozai = ws_1.cell(row=2, column=3).value
        print('所在部门:', suozai)
        # 表2写入所在部门
        ws_2.cell(row=3 + row, column=3).value = suozai

        # 申请部门
        bumen = ws_1.cell(row=3, column=3).value
        print('申请部门:', bumen)
        # 表2写入申请部门
        ws_2.cell(row=3 + row, column=4).value = bumen

        # 申请人
        name = ws_1.cell(row=3, column=11).value
        print('申请人:', name)
        # 表2写入申请人
        ws_2.cell(row=3 + row, column=13).value = name

        # 工务内容
        content = ws_1.cell(row=4, column=3).value
        print('工务内容:', content)
        # 表2写入工务内容
        ws_2.cell(row=3 + row, column=9).value = content
        # 获取表格 填充色 前景颜色
        a = copy(ws_1.cell(row=4, column=3).fill.fgColor.rgb)
        # 判断  公务内容单元格如果有填充色   则在表2  标记黄色
        if a != 'FFFFFFFF':
            # 定义一个前景色
            orange_fill = PatternFill(fill_type='solid', fgColor="FFFF00")
            # 表2更改对应内容的单元格填充色
            ws_2.cell(row=3 + row, column=9).fill = orange_fill
            print('颜色已改....................................................................')
        print('工作内容前景颜色：', type(a), '**', a, '**')

        # 申请日期
        time_new = ws_1.cell(row=5, column=3).value
        print('申请日期:', time_new)
        # 表2写入申请日期
        ws_2.cell(row=3 + row, column=6).value = time_new

        # 要求完成日期
        time_out = ws_1.cell(row=5, column=11).value
        print('要求完成日期:', time_out)
        # 表2写入要求完成日期
        ws_2.cell(row=3 + row, column=7).value = time_out

        # 工务类型
        type_0 = ws_1.cell(row=4, column=3).value
        type_1 = type_0.split('-')[0]
        print('工务类型:', type_1)
        # 表2写入工务类型
        ws_2.cell(row=3 + row, column=5).value = type_1

        # 工务单编号
        num_0 = ws_1.cell(row=1, column=1).value
        num_1 = num_0.split('-')[-1]
        print('工务单编号:', num_1)
        # 表2写入工务单编号
        ws_2.cell(row=3 + row, column=2).value = int(num_1)

        # 打印分割线
        print('*' * 100)

        # 行计数器递增
        row += 1
    # 保存这个表2
    wb_2.save('456.xlsx')


if __name__ == "__main__":
    auto()
