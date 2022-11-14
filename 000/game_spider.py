#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :py_venv -> game_spider.py
# @IDE       :PyCharm
# @Time      :2022/11/14 21:55
# @Author    :杨晓东
# @Email     :lzj155@foxmail.com
# @homepage  :www.demo443.com

import csv
import time
import requests
from lxml import etree
import datetime
import openpyxl


# 装饰器 计时器
def time_out(a_func):
    def clocked(*args, **kwargs):
        start = time.time()
        result = a_func(*args, **kwargs)
        end = time.time()
        print("程序：" + a_func.__name__, "    运行时间：" + str(end - start))
        return result

    return clocked


@time_out
def run():
    # 设置开始网址
    url = "https://www.8591.com.tw/mobileGame-list.html?searchGame=24571&searchServer=26165"
    # 设置代理端口
    proxy = '127.0.0.1:4780'
    proxies = {"http": "http://" + proxy, "https": "http://" + proxy}
    # 设置请求头
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.84 Safari/537.36"}
    # 开始请求网页
    req = requests.get(url, headers=headers, proxies=proxies)
    # 网页编码重置
    req.encoding = 'utf-8'

    # 解析网页
    html = etree.HTML(req.text)
    # 拿到列表数据
    test = html.xpath('//*[@id="content"]/div[2]/div/div[1]/ul/li/a/@title')

    # 获取当前时间
    current_time = datetime.datetime.now()

    # 以当前时间命名文件
    file_name = str(current_time).split(".")[0].replace(':', '').replace(' ', '_')

    # 打印文件名
    print(file_name)

    # 采用xlsx 格式存储数据  方便单元格操作
    # 创建一个xlsx 文件   句柄
    wb = openpyxl.Workbook()  # 创建一个excel文件

    sheet = wb.active  # 获得一个的工作表

    # 表名
    sheet.title = "游戏信息"

    # 分类标识
    a1 = '帳號'
    a2 = '代練'
    a3 = '禮包'
    a4 = '遊戲幣'
    a5 = '代儲'
    a6 = '道具'

    # 修改 表  第一行 为分类信息
    sheet.cell(row=1, column=1).value = a1
    sheet.cell(row=1, column=2).value = a2
    sheet.cell(row=1, column=3).value = a3
    sheet.cell(row=1, column=4).value = a4
    sheet.cell(row=1, column=5).value = a5
    sheet.cell(row=1, column=6).value = a6

    # 行计数器
    r = 1
    # 测试 是否重复
    aaaa = 0

    # 遍历列表数据
    for i in test:
        print(i)
        # 判断  帳號 字段是否存在此次循环
        if a1 in i:
            # 如果在则执行：
            # 行 + 1
            r += 1
            # 循环这一整行
            for j in range(1, r+1):
                # 如果本行 帳號 字段不为空
                if sheet.cell(row=j, column=1).value != None:
                    # 则跳出本次循环
                    continue
                else:
                    # 如果本行 帳號 字段为空
                    # 则将 此次字段内容写入
                    sheet.cell(row=j, column=1).value = i
                    break
                # 以下循环同理
        elif a2 in i:
            r += 1
            for j in range(1, r + 1):
                if sheet.cell(row=j, column=2).value != None:
                    continue
                else:
                    sheet.cell(row=j, column=2).value = i
                    break
        elif a3 in i:
            r += 1
            for j in range(1, r + 1):
                if sheet.cell(row=j, column=3).value != None:
                    continue
                else:
                    sheet.cell(row=j, column=3).value = i
                    break
        elif a4 in i:
            r += 1
            for j in range(1, r + 1):
                if sheet.cell(row=j, column=4).value != None:
                    continue
                else:
                    sheet.cell(row=j, column=4).value = i
                    break
        elif a5 in i:
            r += 1
            for j in range(1, r + 1):
                if sheet.cell(row=j, column=5).value != None:
                    continue
                else:
                    sheet.cell(row=j, column=5).value = i
                    break
        elif a6 in i:
            r += 1
            for j in range(1, r + 1):
                if sheet.cell(row=j, column=6).value != None:
                    continue
                else:
                    sheet.cell(row=j, column=6).value = i
                    break


    # 保存并  以当前时间  重命名文件
    wb.save('./' + file_name + '.xlsx')

    # 打印当前时间
    print('*'*100)
    print(current_time)
    print()

    # ip测试_显示当前ip地址
    ip_url = 'http://icanhazip.com'
    ip_req = requests.get(ip_url, headers=headers, proxies=proxies)
    print('当前ip：', ip_req.text)
    print('*' * 100)


# csv 保存_使用数组方法
def csv_data_save_1():
    # 传入一个文件对象，然后才能在这个文件对象的基础上调用csv的写入方法writerow（写入一行）writerrow（写入多行）

    # 定义标题栏
    headers = ["帳號", "代練", "禮包", "遊戲幣", "代儲", "道具"]
    # 内容_列表（数组）方式
    rows = [
        [1, 'xiaoming', 'male', 168, 23],
        [1, 'xiaohong', 'female', 162, 22],
        [2, 'xiaozhang', 'female', 163, 21],
        [2, 'xiaoli', 'male', 158, 21]
    ]
    # 如果打开csv文件出现空行的情况，那么需要添加一个参数 newline=''
    with open('test.csv', 'w', newline='') as f:
        # 创建csv操作句柄
        f_csv = csv.writer(f)
        # 利用句柄写入标题
        f_csv.writerow(headers)
        # 通过句柄写入内容
        f_csv.writerows(rows)


# csv 保存_使用字典方法
def csv_data_save_2():
    # 写入字典序列类型数据的时候，需要传入两个参数，
    # 一个是文件对象——f，
    # 一个是字段名称——fieldnames，
    # 到时候要写入表头的时候，只需要调用writerheader方法，
    # 写入一行字典系列数据调用writerrow方法，并传入相应字典参数，写入多行调用writerows

    # 定义标题栏
    headers = ['class', 'name', 'sex', 'height', 'year']
    # 内容_字典（序列）方式
    rows = [
        {'class': 1, 'name': 'xiaoming', 'sex': 'male', 'height': 168, 'year': 23},
        {'class': 1, 'name': 'xiaohong', 'sex': 'female', 'height': 162, 'year': 22},
        {'class': 2, 'name': 'xiaozhang', 'sex': 'female', 'height': 163, 'year': 21},
        {'class': 2, 'name': 'xiaoli', 'sex': 'male', 'height': 158, 'year': 21},
    ]
    # 如果打开csv文件出现空行的情况，那么需要添加一个参数 newline=''
    with open('./test.csv', 'w', newline='') as f:
        # 创建csv操作句柄 and 写入标题
        f_csv = csv.DictWriter(f, headers)
        # 调用 writerheader() 写入表头
        f_csv.writeheader()
        # 通过句柄写入内容
        f_csv.writerows(rows)

# csv 读取_方法
def csv_data_read():
    # 读取csv时需要使用reader，并传如一个文件对象，而且reader返回的是一个可迭代的对象，需要使用for循环遍历
    # 在上面，row是一个列表，如果想要查看固定的某列，则需要加上下标，例如我想要查看name，那么只需要改为row[1]
    with open('./test.csv') as f:
        f_csv = csv.reader(f)
        for row in f_csv:
            print(row)
            print(type(row))
            print(row[1])
            print(type(row[1]))

if __name__ == "__main__":
    run()
