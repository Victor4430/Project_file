#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :py_venv -> 工务单统计自动录入.py
# @IDE       :PyCharm
# @Time      :2022/9/8 15:06
# @Author    :杨晓东
# @Email     :lzj155@foxmail.com
# @homepage  :www.demo443.com

import csv
import requests
from lxml import etree
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
import time


# 装饰器 计时器
def time_out(a_func):
    def clocked(*args, **kwargs):
        start = time.time()
        result = a_func(*args, **kwargs)
        end = time.time()
        print("程序：" + a_func.__name__, "    运行时间：" + str(end - start))
        return result

    return clocked


@time_out
def auto():
    # 提示手动输入开始的工作令编号
    start_num = input('请输入结束单号：')
    print('公务单自动录入程序  启动！')
    print('文件过大  请稍等...')

    # 加载公务工作令    表1：
    wb_1 = openpyxl.load_workbook("./2022_10工务工作令（1.1）.xlsx")
    print('正在加载公务工作令.....')

    # 获取所有工作簿名称[]，并打印
    sheets = wb_1.sheetnames

    # 将工作令序号倒序
    sheets = sheets[::-1]

    # 打印出倒序的工作令
    # print(sheets)

    # 加载需要写入的工作簿      表2
    wb_2 = openpyxl.load_workbook('./2022工务单统计（1.0）.xlsx')
    print('正在加载工务单统计.....')

    # 定义表2句柄： 需要写入的工作表
    ws_2 = wb_2['10月份月报_工务单统计']

    # 修改工务单统计表的标题为‘9月份_动力部工务单统计_(8月21-9月20)’
    ws_2.cell(row=1, column=1).value = '10月份动力部月报_(9月21-10月20)'

    # 行计数器
    row = 0
    # 序号计数器
    num = 0

    # 遍历所有的工作表
    for i in sheets:
        # 序号递增
        num += 1

        # 表1句柄： 读取工作表操作：
        ws_1 = wb_1[i]
        print('工作簿编号： ', i)

        # 判断：  跳过  小余  输入的工作表编号
        if int(i) - int(start_num) < 0:
            print('跳过：  ', i)
            # 序号递减
            num -= 1
            # 跳出本次循环
            continue

        # 序号
        print('序号:', num)
        # 表2写入序号
        ws_2.cell(row=3 + row, column=1).value = num

        # 所在部门
        suozai = ws_1.cell(row=2, column=3).value
        print('所在部门:', suozai)
        # 表2写入所在部门
        ws_2.cell(row=3 + row, column=3).value = suozai

        # 申请部门
        bumen = ws_1.cell(row=3, column=3).value
        print('申请部门:', bumen)
        # 表2写入申请部门
        ws_2.cell(row=3 + row, column=4).value = bumen

        # 申请人
        name = ws_1.cell(row=3, column=11).value
        print('申请人:', name)
        # 表2写入申请人
        ws_2.cell(row=3 + row, column=12).value = name

        # 工务内容
        content = ws_1.cell(row=4, column=3).value
        print('工务内容:', content)
        # 表2写入工务内容
        ws_2.cell(row=3 + row, column=9).value = content
        # 获取表格 填充色 前景颜色
        a = copy(ws_1.cell(row=4, column=3).fill.fgColor.rgb)
        # 判断  公务内容单元格如果有填充色   则在表2  标记黄色
        if a != 'FFFFFFFF':
            # 定义一个前景色
            orange_fill = PatternFill(fill_type='solid', fgColor="FFFF00")
            # 表2更改对应内容的单元格填充色
            ws_2.cell(row=3 + row, column=9).fill = orange_fill
            print('颜色已改....................................................................')
        print('工作内容前景颜色：', type(a), '**', a, '**')

        # 申请日期
        time_new = ws_1.cell(row=5, column=3).value
        print('申请日期:', time_new)
        # 表2写入申请日期
        ws_2.cell(row=3 + row, column=6).value = time_new

        # 要求完成日期
        time_out = ws_1.cell(row=5, column=11).value
        print('要求完成日期:', time_out)
        # 表2写入要求完成日期
        ws_2.cell(row=3 + row, column=7).value = time_out

        # 工务类型
        type_0 = ws_1.cell(row=4, column=3).value
        if type_0 == None:
            print('工务类型:', type_0)
            pass
        else:
            type_1 = type_0.split('-')[0]
            print('工务类型:', type_1)
            # 表2写入工务类型
            ws_2.cell(row=3 + row, column=5).value = type_1

        # 工务单编号
        num_0 = ws_1.cell(row=1, column=16).value.replace(' ', '')
        num_1 = num_0.split('-')[-1]
        print('工务单编号:', num_1)
        # 表2写入工务单编号
        ws_2.cell(row=3 + row, column=2).value = int(num_1)

        # 打印分割线
        print('*' * 100)

        # 行计数器递增
        row += 1
    # 保存这个表2
    wb_2.save('./2022工务单统计（1.0）.xlsx')


if __name__ == "__main__":
    auto()
